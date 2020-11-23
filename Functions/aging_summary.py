import pyodbc as db
import pandas as pd
import Functions.operational_functionality as ofn
import xlrd
import sys
from openpyxl import load_workbook

connection = db.connect('DRIVER={SQL Server};'
                        'SERVER=137.116.139.217;'
                        'DATABASE=ARCHIVESKF;'
                        'UID=sa;PWD=erp@123')

def aging_stock_summary_status(name):
    df = pd.read_sql_query("""select * from 

    (select * from 
    (select AGEING as Aging_Days,BRAND,count (distinct T2.ITEMNO) as TotalSKU,itemname , sum(QTYAVAIL) AS TotalStock, T1.AUDTORG  from 
    (select ITEMNO,case when [location] = ('4001') then  'SKF Mirpur Plant'
             when [location] = ('4005') then  'SKF Tongi Plant'
           when [location] = ('4016') then  'SKF Rupganj Plant'
          else AUDTORG end as AUDTORG,AGEING,QTYAVAIL,EXPIRYDATE from ICStockCurrent_Lot WHERE AUDTORG<>'SKFDAT') as T1
    LEFT JOIN
    (select ITEMNO,ITEMNAME,BRAND,PACKSIZE,GPMNAME from PRINFOSKF)as T2
    ON (T1.ITEMNO=T2.ITEMNO)
    WHERE T2.GPMNAME like ? AND AGEING in ('Within 15 Days','Within 30 Days', 'Within 60 Days', 'Within 90 Days', 'Expired')
    Group by BRAND, AGEING,AUDTORG, itemname
    )as T3
    ) t
    pivot
    (sum(TotalStock)
           for AUDTORG in ([BOGSKF],[BSLSKF],[COMSKF],[COXSKF],[CTGSKF],[CTNSKF],[DNJSKF],[FENSKF],[FRDSKF],
    [GZPSKF],[HZJSKF],[JESSKF],[KHLSKF],[KRNSKF],[KSGSKF],[KUSSKF],[MHKSKF],[MIRSKF],[MLVSKF],[MOTSKF],
    [MYMSKF],[NAJSKF],[NOKSKF],[PATSKF],[PBNSKF],[RAJSKF],[RNGSKF],[SAVSKF],[SYLSKF],[TGLSKF],[VRBSKF],[SKF Rupganj Plant],[SKF Mirpur Plant],[SKF Tongi Plant]
    )) AS piv
    
    order by  Aging_Days asc , brand asc
    
         """, connection, params={name})

    # column_values = range(1, len(df))
    # print(column_values)
    # sys.exit()

    writer = pd.ExcelWriter('Data/aging_stock_summary.xlsx', engine='openpyxl')
    wb = writer.book
    df.to_excel(writer, index=False)
    # print('a')
    wb.save('Data/aging_stock_summary.xlsx')
    # print('data')
    cols_aging =[0]
    df_age = pd.read_excel('Data/aging_stock_summary.xlsx', usecols=cols_aging)
    # print(df_age)

    df_age = df_age.replace(['Within 15 Days'], '1 - 15 Days')
    df_age = df_age.replace(['Within 30 Days'], '16 - 30 Days')
    df_age = df_age.replace(['Within 60 Days'], '31 - 60 Days')
    df_age = df_age.replace(['Within 90 Days'], '61 - 90 Days')
    # print(df_age)


    cols_brand = [1]
    df_bran = pd.read_excel('Data/aging_stock_summary.xlsx', usecols=cols_brand)

    cols_all = [2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37]
    df_al = pd.read_excel('Data/aging_stock_summary.xlsx', usecols=cols_all)

    writer = pd.ExcelWriter('Data/aging_stock_summary_copy.xlsx', engine='xlsxwriter')
    df_age.to_excel(writer, sheet_name='Sheet1', index=False, startcol=0, startrow=0)
    df_bran.to_excel(writer, sheet_name='Sheet1', index=False, startcol=1, startrow=0)
    df_al.to_excel(writer, sheet_name='Sheet1', index=False, startcol=2, startrow=0)

    # worksheet = writer.sheets['Sheet1']
    # worksheet.set_column('D:D', 40)
    # worksheet.set_column('F:AZ', 20)
    # worksheet.set_column('BA:CE', 15)
    # worksheet.set_column('CF:CJ', 20)

    writer.save()

    # sys.exit()
    #
    column_values = range(1,len(df)+1)
    # df_new = pd.DataFrame({'BSL NO': column_values})
    df_new = pd.DataFrame({'ISL NO': column_values})
    wb = load_workbook('Data/aging_stock_summary_copy.xlsx')

    ws = wb['Sheet1']

    # for index, row in df_new.iterrows():
    #     cell = 'B%d' % (index + 2)
    #     ws[cell] = row[0]
    #
    # ws.cell(row=1, column=2).value = 'BSL NO'

    for index, row in df_new.iterrows():
        cell = 'C%d' % (index + 2)
        ws[cell] = row[0]

    ws.cell(row=1, column=3).value = 'ISL NO'

    wb.save('Data/aging_stock_summary_copy.xlsx')
    #
    print('11. Branch - Brand - SKU wise Stock Aging Status Dataaved')


    def check(val):
        if val =='':
            return '-'
        else:
            return int(val)
    # sys.exit()
    excel_data_df = pd.read_excel('Data/aging_stock_summary_copy.xlsx', sheet_name='Sheet1',
                                  usecols=['Aging_Days', 'BRAND'])
    aging_days = ofn.create_dup_count_list(excel_data_df, 'Aging_Days')
    brand = ofn.create_dup_count_list(excel_data_df, 'BRAND')

    wb = xlrd.open_workbook('Data/aging_stock_summary_copy.xlsx')
    sh = wb.sheet_by_name('Sheet1')

    tabletd = ""

    for i in range(1, sh.nrows):
        tabletd = tabletd + "<tr>\n"

        for j in range(0, 1):
            # BSL NO
            if (aging_days[i - 1] != 0):
                tabletd = tabletd + "<td class=\"serial\" rowspan=\"" + str((aging_days[i - 1])) + "\"> " + str(
                    sh.cell_value(i, j)) + "</td>\n"

        for j in range(1, 2):
            # Brand
            if (brand[i - 1] != 0):
                tabletd = tabletd + "<td class=\"brandtd\" rowspan=\"" + str(brand[i - 1]) + "\">" + str(
                    sh.cell_value(i, j)) + "</td>\n"

        for j in range(2, 3):
            # ITemNo
            tabletd = tabletd + "<td class=\"sl_center\">" + str(int(sh.cell_value(i, j))) + "</td>\n"

        for j in range(3, 4):
            # ITemNo
            tabletd = tabletd + "<td class=\"y_desc_sales\">" + str((sh.cell_value(i, j))) + "</td>\n"

        for j in range(4, 5):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(5, 6):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(6, 7):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(7, 8):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(8, 9):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(9, 10):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(10, 11):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(11, 12):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(12, 13):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(13, 14):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(14, 15):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(15, 16):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(16, 17):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(17, 18):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(18, 19):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(19, 20):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(20, 21):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(21, 22):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(22, 23):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(23, 24):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(24, 25):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(25, 26):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(26, 27):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(27, 28):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(28, 29):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(29, 30):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(30, 31):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(31, 32):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(32, 33):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(33, 34):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(34, 35):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(35, 36):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(36, 37):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        for j in range(37, 38):
            tabletd = tabletd + "<td class=\"number_style\">" + str((check(sh.cell_value(i, j)))) + "</td>\n"

        table = tabletd + "</tr>\n"
    print("11.1. Branch - Brand - SKU wise Stock Aging Status: Summary Table created \n")
    return table
