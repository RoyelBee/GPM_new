import pyodbc as db
import pandas as pd
import Functions.operational_functionality as ofn
import xlrd
import sys
from openpyxl import load_workbook

connection = db.connect('DRIVER={SQL Server};'
                        'SERVER=137.116.139.217;'
                        'DATABASE=ARCHIVESKF;'
                        'UID=sa;PWD=erp@123')

def get_branch_aging_stock_status(name):
    df = pd.read_sql_query(""" Select * from
        (select *
        from
        (
        Select * from
        (select dense_rank() OVER (ORDER BY BRAND  ) [BSL NO], BRAND,(select ROW_NUMBER() OVER (ORDER BY (SELECT NULL))) AS [ISL NO],ITEMNAME,AUDTORG, Max(AGEING) AS AGEING   from
        (select ITEMNO,
        case when [location] = ('4001') then  'SKF Mirpur Plant'
        when [location] = ('4005') then  'SKF Tongi Plant'
        when [location] = ('4016') then  'SKF Rupganj Plant'
        else AUDTORG end as AUDTORG,AGEING,QTYAVAIL,EXPIRYDATE from ICStockCurrent_Lot) as T1
        LEFT JOIN
        (select ITEMNO,ITEMNAME,BRAND,PACKSIZE,GPMNAME from PRINFOSKF )as T2
        ON (T1.ITEMNO=T2.ITEMNO)
        WHERE T2.GPMNAME like ?
        Group by  BRAND,ITEMNAME,AUDTORG ) as T3
        ) src
        pivot
        (Max(AGEING)
        for AUDTORG in ([BOGSKF],[BSLSKF],[COMSKF],[COXSKF],[CTGSKF],[CTNSKF],[DNJSKF],[FENSKF],[FRDSKF],[GZPSKF],[HZJSKF],[JESSKF],[KHLSKF],[KRNSKF],[KSGSKF],[KUSSKF],[MHKSKF],[MIRSKF],[MLVSKF],[MOTSKF],[MYMSKF],[NAJSKF],[NOKSKF],[PATSKF],[PBNSKF],[RAJSKF],[RNGSKF],[SAVSKF],[SKFDAT],[SYLSKF],[TGLSKF],[VRBSKF],[SKF Rupganj Plant],[SKF Mirpur Plant],[SKF Tongi Plant]
        )) AS piv
        ) as TblAging
        left join
        (select *
        from
        (Select * from
        (select dense_rank() OVER (ORDER BY BRAND  ) [BSL NO], BRAND,ITEMNAME,AUDTORG, Sum(QTYAVAIL) AS TotalStock   from
        (select ITEMNO,case when [location] = ('4001') then  'SKF Mirpur Plant'
        when [location] = ('4005') then  'SKF Tongi Plant'
        when [location] = ('4016') then  'SKF Rupganj Plant'
        else AUDTORG end as AUDTORG,AGEING,QTYAVAIL,EXPIRYDATE from ICStockCurrent_Lot) as T1
        LEFT JOIN
        (select ITEMNO,ITEMNAME,BRAND,PACKSIZE,GPMNAME from PRINFOSKF )as T2
        ON (T1.ITEMNO=T2.ITEMNO)
        WHERE T2.GPMNAME like ?
        Group by  BRAND,ITEMNAME,AUDTORG ) as T3
        ) src
        pivot
        (
        sum(TotalStock)
        for AUDTORG in ([BOGSKF],[BSLSKF],[COMSKF],[COXSKF],[CTGSKF],[CTNSKF],[DNJSKF],[FENSKF],[FRDSKF],[GZPSKF],[HZJSKF],[JESSKF],[KHLSKF],[KRNSKF],[KSGSKF],[KUSSKF],[MHKSKF],[MIRSKF],[MLVSKF],[MOTSKF],[MYMSKF],[NAJSKF],[NOKSKF],[PATSKF],[PBNSKF],[RAJSKF],[RNGSKF],[SAVSKF],[SKFDAT],[SYLSKF],[TGLSKF],[VRBSKF],[SKF Rupganj Plant],[SKF Mirpur Plant],[SKF Tongi Plant]
        )
        ) AS piv) as tblsStock
        on (tblsStock.ITEMNAME=TblAging.ITEMNAME)
        
             """, connection, params=(name, name))

    # column_values = range(1, len(df))
    # print(column_values)
    # sys.exit()

    writer = pd.ExcelWriter('Data/branch_wise_aging_stock.xlsx', engine='openpyxl')
    wb = writer.book
    df.to_excel(writer, index=False)
    wb.save('Data/branch_wise_aging_stock.xlsx')

    column_values = range(1,len(df)+1)
    df_new = pd.DataFrame({'ISL NO': column_values})
    wb = load_workbook('Data/branch_wise_aging_stock.xlsx')

    ws = wb['Sheet1']

    for index, row in df_new.iterrows():
        cell = 'C%d' % (index + 2)
        ws[cell] = row[0]
    wb.save('Data/branch_wise_aging_stock.xlsx')

    print('17. Branch - Brand - SKU wise Stock Aging Status: Detailed data saved')

    # sys.exit()
    excel_data_df = pd.read_excel('Data/branch_wise_aging_stock.xlsx', sheet_name='Sheet1',
                                  usecols=['BSL NO', 'BRAND'])
    bslno = ofn.create_dup_count_list(excel_data_df, 'BSL NO')
    brand = ofn.create_dup_count_list(excel_data_df, 'BRAND')

    wb = xlrd.open_workbook('Data/branch_wise_aging_stock.xlsx')
    sh = wb.sheet_by_name('Sheet1')

    tabletd = ""

    for i in range(1, sh.nrows):
        tabletd = tabletd + "<tr>\n"

        for j in range(0, 1):
            # BSL NO
            if (bslno[i - 1] != 0):
                tabletd = tabletd + "<td class=\"serial\" rowspan=\"" + str((bslno[i - 1])) + "\"> " + str(
                    int(sh.cell_value(i, j))) + "</td>\n"

        for j in range(1, 2):
            # Brand
            if (brand[i - 1] != 0):
                tabletd = tabletd + "<td class=\"brandtd\" rowspan=\"" + str(brand[i - 1]) + "\">" + str(
                    sh.cell_value(i, j)) + "</td>\n"

        for j in range(2, 3):
            # ITemNo
            tabletd = tabletd + "<td class=\"serial\">" + str(int(sh.cell_value(i, j))) + "</td>\n"

        for j in range(3, 4):
            # ITemNo
            tabletd = tabletd + "<td class=\"y_desc_sales\">" + str((sh.cell_value(i, j))) + "</td>\n"

        for j in range(4, 5):
            bog = sh.cell_value(i, j)

        for j in range(5, 6):
            bsl = sh.cell_value(i, j)

        for j in range(6, 7):
            com = sh.cell_value(i, j)

        for j in range(7, 8):
            cox = sh.cell_value(i, j)

        for j in range(8, 9):
            ctg = sh.cell_value(i, j)

        for j in range(9, 10):
            ctn = sh.cell_value(i, j)

        for j in range(10, 11):
            dnj = sh.cell_value(i, j)

        for j in range(11, 12):
            fen = sh.cell_value(i, j)

        for j in range(12, 13):
            frd = sh.cell_value(i, j)

        for j in range(13, 14):
            gzp = sh.cell_value(i, j)

        for j in range(14, 15):
            hzj = sh.cell_value(i, j)

        for j in range(15, 16):
            jes = sh.cell_value(i, j)

        for j in range(16, 17):
            khl = sh.cell_value(i, j)

        for j in range(17, 18):
            krn = sh.cell_value(i, j)

        for j in range(18, 19):
            ksg = sh.cell_value(i, j)

        for j in range(19, 20):
            kus = sh.cell_value(i, j)

        for j in range(20, 21):
            mhk = sh.cell_value(i, j)

        for j in range(21, 22):
            mir = sh.cell_value(i, j)

        for j in range(22, 23):
            mlv = sh.cell_value(i, j)

        for j in range(23, 24):
            mot = sh.cell_value(i, j)

        for j in range(24, 25):
            mym = sh.cell_value(i, j)

        for j in range(25, 26):
            naj = sh.cell_value(i, j)

        for j in range(26, 27):
            nok = sh.cell_value(i, j)

        for j in range(27, 28):
            pat = sh.cell_value(i, j)

        for j in range(28, 29):
            pbn = sh.cell_value(i, j)

        for j in range(29, 30):
            raj = sh.cell_value(i, j)

        for j in range(30, 31):
            rng = sh.cell_value(i, j)

        for j in range(31, 32):
            sav = sh.cell_value(i, j)

        for j in range(33, 34):
            syl = sh.cell_value(i, j)

        for j in range(34, 35):
            tgl = sh.cell_value(i, j)

        for j in range(35, 36):
            vrb = sh.cell_value(i, j)

        for j in range(36, 37):
            Rupganj = sh.cell_value(i, j)

        for j in range(37, 38):
            Mirpur = sh.cell_value(i, j)

        for j in range(38, 39):
            Tongi = sh.cell_value(i, j)


        for j in range(42, 43):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(bog)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(43, 44):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(bsl)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(44, 45):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(com)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(45, 46):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(cox)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(46, 47):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(ctg)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(47, 48):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(ctn)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(48, 49):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(dnj)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(49, 50):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(fen)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(50, 51):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(frd)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(51, 52):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(gzp)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(52, 53):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(hzj)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(53, 54):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(jes)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(54, 55):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(khl)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(55, 56):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(krn)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(56, 57):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(ksg)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(57, 58):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(kus)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(58, 59):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(mhk)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(59, 60):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(mir)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(60, 61):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(mlv)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(61, 62):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(mot)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(62, 63):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(mym)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(63, 64):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(naj)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(64, 65):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(nok)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(65, 66):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(pat)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(66, 67):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(pbn)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(67, 68):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(raj)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(68, 69):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(rng)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(69, 70):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(sav)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"



        for j in range(71, 72):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(syl)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(72, 73):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(tgl)) + "\">"+ str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(73, 74):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(vrb)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(74, 75):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(Rupganj)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(75, 76):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(Mirpur)) + "\">"+ str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"

        for j in range(76, 77):
            tabletd = tabletd + "<td class=\"number_style\"style=\"background-color:" + \
                      str(ofn.status_color(Tongi)) + "\">" + str(ofn.integer_converter(sh.cell_value(i, j))) + "</td>\n"


        table = tabletd + "</tr>\n"
    print("17.1. Branch - Brand - SKU wise Stock Aging Status: Detailed table Created\n")
    return table
